from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash
)

from flask_login import (
    LoginManager,
    login_user,
    login_required,
    logout_user,
    current_user
)

from models import (
    db,
    User,
    Exam,
    Question,
    QuestionBank,
    ExamSession
)

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)
from flask_migrate import Migrate
from cbt_engine import CBTSessionEngine
from flask_mail import Mail, Message
from datetime import datetime, UTC
from io import BytesIO
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from itsdangerous import URLSafeTimedSerializer
from flask_mail import Message
import uuid
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
from docx import Document
import random
import json
import os

from models import (
    db,
    User,
    Exam,
    Question,
    ExamSession
)

app = Flask(__name__)

# ======================
# GRADE ENGINE (HELPER)
# ======================

def calculate_grade(score, total):
    percent = (score / total) * 100

    if percent >= 70:
        return "A"
    elif percent >= 60:
        return "B"
    elif percent >= 50:
        return "C"
    elif percent >= 45:
        return "D"
    else:
        return "F"
        
# ======================
# CONFIG
# ======================
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "shalomsp100@gmail.com"
app.config["MAIL_PASSWORD"] = "tqwubutzasovybrm"

app.config["MAIL_DEFAULT_SENDER"] = "shalomsp100@gmail.com"
app.config["MAIL_MAX_EMAILS"] = None
app.config["MAIL_SUPPRESS_SEND"] = False

mail = Mail(app)

app.config["SECRET_KEY"] = "CHANGE_TO_SECURE_KEY"
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
serializer = URLSafeTimedSerializer(
    app.config["SECRET_KEY"]
)
db.init_app(app)
migrate = Migrate(app, db)

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
import uuid


def generate_certificate(user, exam, score, total):

    cert_id = str(uuid.uuid4())[:8].upper()
    grade = calculate_grade(score, total)

    file_path = f"certificates/{cert_id}.pdf"

    os.makedirs("certificates", exist_ok=True)

    c = canvas.Canvas(file_path, pagesize=letter)

    c.setFont("Helvetica-Bold", 20)
    c.drawString(180, 750, "SHALOM NEXUS ACADEMY")

    c.setFont("Helvetica", 14)
    c.drawString(200, 700, "CERTIFICATE OF COMPLETION")

    c.setFont("Helvetica", 12)
    c.drawString(100, 650, f"This certifies that: {user.full_name}")
    c.drawString(100, 630, f"Has completed: {exam.title}")
    c.drawString(100, 610, f"Score: {score}/{total}")
    c.drawString(100, 590, f"Grade: {grade}")
    c.drawString(100, 570, f"Certificate ID: {cert_id}")

    c.drawString(100, 520, "Congratulations on your achievement!")

    c.save()

    return file_path, cert_id, Grade

def send_certificate_email(user, exam, file_path):

    try:
        msg = Message(
            subject="Your Certificate - Shalom Nexus",
            sender=app.config["MAIL_USERNAME"],
            recipients=[user.email]
        )

        msg.body = f"""
Hello {user.full_name},

Congratulations!

You have successfully completed {exam.title}.

Your certificate is attached.

Regards,
Shalom Nexus Academy
"""

        with app.open_resource(file_path) as fp:
            msg.attach("certificate.pdf", "application/pdf", fp.read())

        mail.send(msg)

    except Exception as e:
        print("CERT EMAIL ERROR:", e)

from flask_mail import Message

def send_welcome_email(user):

    msg = Message(
        subject="Welcome to Shalom Nexus CBT",
        sender=app.config["MAIL_USERNAME"],
        recipients=[user.email]
    )

    msg.body = f"""
Hello {user.full_name},

Welcome to Shalom Nexus Academy.

Your account has been successfully created.

Name: {user.full_name}
Email: {user.email}

You may now log in and access:
• CBT Examinations
• Results
• Certificates
• Performance Reports

Regards,
Shalom Nexus CBT Team
"""

    mail.send(msg)
        
with app.app_context():
    db.create_all()

    admin = User.query.filter_by(
        email="odeleyesamuel2020@gmail.com"
    ).first()

    # =========================
    # 🔥 FORCE ADMIN CORRECTION
    # =========================
    if not admin:

        admin = User(
            full_name="System Administrator",
            email="odeleyesamuel2020@gmail.com",
            password=generate_password_hash("admin123"),
            role="admin",
            registration_number="ADMIN001"
        )

        db.session.add(admin)
        db.session.commit()

        print("✅ Default admin created")

    else:
        # 🔥 FIX EXISTING ADMIN IF BROKEN
        changed = False

        if admin.role != "admin":
            admin.role = "admin"
            changed = True

        if not admin.registration_number:
            admin.registration_number = "ADMIN001"
            changed = True

        if changed:
            db.session.commit()
            print("🔧 Existing admin fixed")
            
# ======================
# LOGIN SETUP
# ======================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(
        User,
        int(user_id)
    )


# ======================
# ROLE GUARD
# ======================
def admin_required(func):
    def wrapper(*args, **kwargs):
        if current_user.role != "admin":
            return redirect(url_for("dashboard"))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


# ======================
# HOME
# ======================
@app.route("/")
def index():

    students_count = User.query.filter_by(role="student").count()
    exams_count = Exam.query.count()
    questions_count = Question.query.count() + QuestionBank.query.count()

    recent_exams = Exam.query.filter_by(is_published=True)\
        .order_by(Exam.created_at.desc())\
        .limit(6).all()

    top_students = User.query.filter_by(role="student")\
        .order_by(User.created_at.desc())\
        .limit(8).all()

    # =========================
    # 🔴 LIVE ACTIVITY SYSTEM
    # =========================
    actions = [
        "completed a CBT exam",
        "scored above 80% in Mathematics",
        "started WAEC English test",
        "joined Shalom Nexus Academy",
        "completed NECO Physics practice",
        "improved ranking on leaderboard"
    ]

    users = User.query.filter_by(role="student").limit(10).all()

    live_feed = []

    for _ in range(10):
        if users:
            u = random.choice(users)
            live_feed.append({
                "name": u.full_name,
                "action": random.choice(actions),
                "time": f"{random.randint(1, 59)} mins ago"
            })

    # =========================
    # 🟢 FAKE ONLINE COUNT (REALISTIC)
    # =========================
    online_count = 120 + random.randint(0, 80)

    # =========================
    # 🟡 SAFE FALLBACKS
    # =========================
    if not recent_exams:
        recent_exams = []

    if not top_students:
        top_students = []

    return render_template(
        "index.html",
        students_count=students_count,
        exams_count=exams_count,
        questions_count=questions_count,
        recent_exams=recent_exams,
        top_students=top_students,
        live_feed=live_feed,
        online_count=online_count
    )

# ======================
# AUTH
# ======================
@app.route("/register", methods=["GET", "POST"])
def register():

    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":

        full_name = request.form.get(
            "full_name", ""
        ).strip()

        email = request.form.get(
            "email", ""
        ).strip().lower()

        phone_number = request.form.get(
            "phone_number", ""
        ).strip()

        institution = request.form.get(
            "institution", ""
        ).strip()

        gender = request.form.get(
            "gender", ""
        ).strip()

        password = request.form.get(
            "password", ""
        )

        confirm_password = request.form.get(
            "confirm_password", ""
        )

        # =====================
        # VALIDATION
        # =====================

        if not full_name:
            flash(
                "Full name is required.",
                "danger"
            )
            return redirect(url_for("register"))

        if len(full_name) < 3:
            flash(
                "Full name must be at least 3 characters.",
                "danger"
            )
            return redirect(url_for("register"))

        if not email:
            flash(
                "Email address is required.",
                "danger"
            )
            return redirect(url_for("register"))

        if User.query.filter_by(
            email=email
        ).first():

            flash(
                "Email already exists.",
                "warning"
            )
            return redirect(url_for("register"))

        if len(password) < 6:
            flash(
                "Password must be at least 6 characters.",
                "danger"
            )
            return redirect(url_for("register"))

        if password != confirm_password:

            flash(
                "Passwords do not match.",
                "danger"
            )
            return redirect(url_for("register"))

        # =====================
        # REGISTRATION NUMBER
        # =====================

        next_id = User.query.count() + 1

        registration_number = (
            f"SNX{datetime.now().year}"
            f"{next_id:05d}"
        )

        hashed_password = generate_password_hash(
            password
        )

        user = User(
            full_name=full_name,
            email=email,
            password=hashed_password,
            role="student",
            registration_number=registration_number,
            phone_number=phone_number,
            institution=institution,
            gender=gender
        )

        db.session.add(user)
        db.session.commit()

        try:
            send_welcome_email(user)
        except Exception as e:
            print(e)

        flash(
            "Registration successful. Please login.",
            "success"
        )

        return redirect(
            url_for("login")
        )

    return render_template(
        "register.html"
    )

#login 
@app.route("/login", methods=["GET", "POST"])
def login():

    if current_user.is_authenticated:

        if current_user.role == "admin":
            return redirect(
                url_for("admin_dashboard")
            )

        return redirect(
            url_for("dashboard")
        )

    if request.method == "POST":

        identifier = request.form.get(
            "identifier",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        user = User.query.filter(
            (User.email == identifier) |
            (User.registration_number == identifier)
        ).first()

        if not user:
            return render_template(
                "login.html",
                error="Invalid login details"
            )

        if not check_password_hash(
            user.password,
            password
        ):
            return render_template(
                "login.html",
                error="Invalid login details"
            )

        login_user(user)

        flash(
            f"Welcome back, {user.full_name}",
            "success"
        )

        if user.role == "admin":
            return redirect(
                url_for("admin_dashboard")
            )

        return redirect(
            url_for("dashboard")
        )

    return render_template(
        "login.html"
    )

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))

#Admin for student management
@app.route("/admin/users")
@login_required
def manage_users():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    search = request.args.get("search", "").strip()

    query = User.query.filter_by(role="student")

    if search:
        query = query.filter(
            (User.full_name.ilike(f"%{search}%")) |
            (User.email.ilike(f"%{search}%")) |
            (User.registration_number.ilike(f"%{search}%"))
        )

    students = query.order_by(User.created_at.desc()).all()

    total_students = User.query.filter_by(role="student").count()

    return render_template(
        "manage_users.html",
        students=students,
        total_students=total_students,
        search=search
    )

#student profile
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():

    if request.method == "POST":

        current_user.full_name = request.form.get("full_name")
        current_user.phone_number = request.form.get("phone_number")
        current_user.institution = request.form.get("institution")
        current_user.gender = request.form.get("gender")

        db.session.commit()

        flash("Profile updated successfully", "success")

        return redirect(url_for("profile"))

    return render_template("profile.html", user=current_user)

#FORGET PASSWORD
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        email = request.form.get("email")

        user = User.query.filter_by(
            email=email
        ).first()

        if user:

            token = serializer.dumps(
                user.email,
                salt="password-reset"
            )

            reset_link = url_for(
                "reset_password",
                token=token,
                _external=True
            )

            msg = Message(
                subject="Shalom Nexus CBT Password Reset",
                recipients=[user.email],
                sender=app.config["MAIL_USERNAME"]
            )

            msg.html = f"""
            <h2>Shalom Nexus CBT</h2>

            <p>Hello {user.full_name},</p>

            <p>
            A password reset request was received
            for your account.
            </p>

            <hr>

            <p>
            <strong>Registration Number:</strong>
            {user.registration_number}
            </p>

            <p>
            <strong>Email:</strong>
            {user.email}
            </p>

            <hr>

            <p>
            Click below to reset your password:
            </p>

            <p>
            <a href="{reset_link}">
                Reset Password
            </a>
            </p>

            <p>
            This link expires in 30 minutes.
            </p>

            <p>
            If you did not request this,
            simply ignore this email.
            </p>
            """

            mail.send(msg)

        flash(
            "If the email exists, a reset link has been sent.",
            "success"
        )

        return redirect(
            url_for("login")
        )

    return render_template(
        "forgot_password.html"
    )
    

@app.route(
    "/reset-password/<token>",
    methods=["GET", "POST"]
)
def reset_password(token):

    try:
        email = serializer.loads(
            token,
            salt="password-reset",
            max_age=1800
        )

    except Exception:
        flash(
            "Invalid or expired reset link.",
            "danger"
        )

        return redirect(
            url_for("login")
        )

    user = User.query.filter_by(
        email=email
    ).first_or_404()

    if request.method == "POST":

        password = request.form.get("password")

        confirm_password = request.form.get(
            "confirm_password"
        )

        if password != confirm_password:

            flash(
                "Passwords do not match.",
                "danger"
            )

            return redirect(request.url)

        user.password = generate_password_hash(
            password
        )

        db.session.commit()

        flash(
            "Password updated successfully.",
            "success"
        )

        return redirect(
            url_for("login")
        )

    return render_template(
        "reset_password.html"
    )
    
# ======================
# DASHBOARDS
# ======================
@app.route("/dashboard")
@login_required
def dashboard():

    if current_user.role == "admin":
        return redirect(url_for("admin_dashboard"))

    now = datetime.utcnow()

    # ACTIVE EXAMS ONLY
    active_exams = Exam.query.filter(
        Exam.is_published == True,
        (Exam.start_date == None) | (Exam.start_date <= now),
        (Exam.end_date == None) | (Exam.end_date >= now)
    ).order_by(Exam.id.desc()).all()

    # USER ATTEMPTS
    attempts = ExamSession.query.filter_by(
        user_id=current_user.id
    ).order_by(ExamSession.start_time.desc()).limit(10).all()

    # STATS
    total_exams = len(active_exams)
    total_attempts = ExamSession.query.filter_by(
        user_id=current_user.id
    ).count()

    avg_score = db.session.query(
        db.func.avg(ExamSession.score)
    ).filter_by(
        user_id=current_user.id
    ).scalar() or 0

    return render_template(
        "dashboard.html",
        exams=active_exams,
        attempts=attempts,
        total_exams=total_exams,
        total_attempts=total_attempts,
        avg_score=round(avg_score, 2)
    )


@app.route("/admin/dashboard")
@login_required
def admin_dashboard():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    now = datetime.utcnow()

    # BASIC STATS
    students = User.query.filter_by(role="student").count()
    exams = Exam.query.count()
    questions = Question.query.count()
    attempts = ExamSession.query.count()

    # EXAM BREAKDOWN (IMPORTANT UPGRADE)
    published_exams = Exam.query.filter_by(is_published=True).count()

    draft_exams = Exam.query.filter_by(is_published=False).count()

    active_exams = Exam.query.filter(
        Exam.is_published == True,
        (Exam.start_date == None) | (Exam.start_date <= now),
        (Exam.end_date == None) | (Exam.end_date >= now)
    ).count()

    # RECENT ACTIVITY (LAST 10 ATTEMPTS)
    recent_attempts = ExamSession.query.order_by(
        ExamSession.start_time.desc()
    ).limit(10).all()

    # TOP EXAMS (BY QUESTIONS COUNT)
    top_exams = db.session.query(
        Exam,
        db.func.count(Question.id).label("q_count")
    ).join(
        Question, Question.exam_id == Exam.id
    ).group_by(
        Exam.id
    ).order_by(
        db.desc("q_count")
    ).limit(5).all()

    return render_template(
        "admin_dashboard.html",
        user=current_user,
        students=students,
        exams=exams,
        questions=questions,
        attempts=attempts,
        published_exams=published_exams,
        draft_exams=draft_exams,
        active_exams=active_exams,
        recent_attempts=recent_attempts,
        top_exams=top_exams
    )


# ======================
# CREATE EXAM
# ======================
@app.route("/admin/create-exam", methods=["GET", "POST"])
@login_required
def create_exam():

    if current_user.role != "admin":
        return redirect(url_for("exams"))

    if request.method == "POST":

        exam = Exam(
            title=request.form.get("title"),

            subject=request.form.get("subject"),

            description=request.form.get(
                "description"
            ),

            duration=int(
                request.form.get(
                    "duration",
                    60
                )
            ),

            pass_mark=int(
                request.form.get(
                    "pass_mark",
                    50
                )
            ),

            instructions=request.form.get(
                "instructions"
            ),

            shuffle_questions=(
                request.form.get(
                    "shuffle_questions"
                ) == "yes"
            ),

            shuffle_options=(
                request.form.get(
                    "shuffle_options"
                ) == "yes"
            ),

            allow_review=(
                request.form.get(
                    "allow_review"
                ) == "yes"
            ),

            allow_multiple_attempts=(
                request.form.get(
                    "allow_multiple_attempts"
                ) == "yes"
            ),

            is_published=(
                request.form.get(
                    "is_published"
                ) == "1"
            ),

            created_by=current_user.id
        )

        db.session.add(exam)
        db.session.commit()

        flash(
            "Exam created successfully",
            "success"
        )

        return redirect(
            url_for(
                "manage_exam",
                exam_id=exam.id
            )
        )

    return render_template(
        "create_exam.html"
    )

# MANAGE EXAM
@app.route("/admin/exam/<int:exam_id>/manage")
@login_required
def manage_exam(exam_id):

    exam = Exam.query.get_or_404(exam_id)

    questions = Question.query.filter_by(
        exam_id=exam_id
    ).all()

    print("Exam:", exam.title)
    print("Questions found:", len(questions))

    for q in questions:
        print(q.id, q.question, q.status)

    return render_template(
        "manage_exam.html",
        exam=exam,
        questions=questions
    )


@app.route("/admin/exams")
@login_required
def admin_exams():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    exams = Exam.query.all()

    return render_template("admin_exams.html", exams=exams)
    
    
@app.route(
    "/admin/question-builder/<int:exam_id>",
    methods=["GET", "POST"]
)
@login_required
def question_builder(exam_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    exam = Exam.query.get_or_404(exam_id)

    if request.method == "POST":

        questions = request.form.getlist("question[]")
        option_a = request.form.getlist("option_a[]")
        option_b = request.form.getlist("option_b[]")
        option_c = request.form.getlist("option_c[]")
        option_d = request.form.getlist("option_d[]")
        answers = request.form.getlist("answer[]")

        saved_questions = []

        for q, a, b, c, d, ans in zip(
            questions,
            option_a,
            option_b,
            option_c,
            option_d,
            answers
        ):

            if not q or not q.strip():
                continue

            # Save to Exam Questions
            new_question = Question(
                exam_id=exam.id,
                question=q.strip(),
                option_a=a.strip() if a else "",
                option_b=b.strip() if b else "",
                option_c=c.strip() if c else "",
                option_d=d.strip() if d else "",
                answer=ans.strip().upper(),
                marks=1
            )

            db.session.add(new_question)

            # Save to Question Bank
            bank_question = QuestionBank(
                subject=exam.subject,
                topic="General",
                question=q.strip(),
                option_a=a.strip() if a else "",
                option_b=b.strip() if b else "",
                option_c=c.strip() if c else "",
                option_d=d.strip() if d else "",
                answer=ans.strip().upper()
            )

            db.session.add(bank_question)

            saved_questions.append(new_question)

        db.session.commit()

        flash(
            f"{len(saved_questions)} question(s) saved successfully!",
            "success"
        )

        return redirect(
            url_for(
                "manage_exam",
                exam_id=exam.id
            )
        )

    return render_template(
        "question_builder.html",
        exam=exam
    )
    
# ======================
# PUBLISH EXAM
# ======================
@app.route("/admin/publish-exam/<int:exam_id>")
@login_required
def publish_exam(exam_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    exam = Exam.query.get_or_404(exam_id)

    exam.is_published = True
    db.session.commit()

    flash("Exam published successfully!", "success")

    return redirect(url_for("manage_exam", exam_id=exam.id))


@app.route(
    "/admin/question/delete/<int:question_id>"
)
@login_required
def delete_question(question_id):

    question = Question.query.get_or_404(
        question_id
    )

    exam_id = question.exam_id

    db.session.delete(question)

    db.session.commit()

    flash(
        "Question deleted successfully",
        "success"
    )

    return redirect(
        url_for(
            "manage_exam",
            exam_id=exam_id
        )
    )
    
@app.route(
    "/admin/exam/duplicate/<int:exam_id>"
)
@login_required
def duplicate_exam(exam_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    old_exam = Exam.query.get_or_404(exam_id)

    new_exam = Exam(
        title=f"{old_exam.title} (Copy)",
        subject=old_exam.subject,
        description=old_exam.description,
        duration=old_exam.duration,
        pass_mark=old_exam.pass_mark,
        instructions=old_exam.instructions,
        shuffle_questions=old_exam.shuffle_questions,
        shuffle_options=old_exam.shuffle_options,
        allow_review=old_exam.allow_review,
        allow_multiple_attempts=old_exam.allow_multiple_attempts,
        is_published=False,
        created_by=current_user.id
    )

    db.session.add(new_exam)
    db.session.commit()

    questions = Question.query.filter_by(
        exam_id=old_exam.id
    ).all()

    for q in questions:

        clone = Question(
            exam_id=new_exam.id,
            question=q.question,
            option_a=q.option_a,
            option_b=q.option_b,
            option_c=q.option_c,
            option_d=q.option_d,
            answer=q.answer,
            marks=q.marks
        )

        db.session.add(clone)

    db.session.commit()

    flash(
        "Exam duplicated successfully",
        "success"
    )

    return redirect(
        url_for(
            "manage_exam",
            exam_id=new_exam.id
        )
    )
    

@app.route(
    "/admin/exam/edit/<int:exam_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_exam(exam_id):

    if current_user.role != "admin":
        return redirect(
            url_for("dashboard")
        )

    exam = Exam.query.get_or_404(
        exam_id
    )

    if request.method == "POST":

        exam.title = request.form.get(
            "title"
        )

        exam.subject = request.form.get(
            "subject"
        )

        exam.description = request.form.get(
            "description"
        )

        exam.duration = int(
            request.form.get(
                "duration",
                60
            )
        )

        exam.pass_mark = int(
            request.form.get(
                "pass_mark",
                50
            )
        )

        exam.instructions = request.form.get(
            "instructions"
        )

        exam.shuffle_questions = (
            request.form.get(
                "shuffle_questions"
            ) == "yes"
        )

        exam.shuffle_options = (
            request.form.get(
                "shuffle_options"
            ) == "yes"
        )

        exam.allow_review = (
            request.form.get(
                "allow_review"
            ) == "yes"
        )

        exam.allow_multiple_attempts = (
            request.form.get(
                "allow_multiple_attempts"
            ) == "yes"
        )

        exam.is_published = (
            request.form.get(
                "is_published"
            ) == "1"
        )

        db.session.commit()

        flash(
            "Exam updated successfully",
            "success"
        )

        return redirect(
            url_for(
                "manage_exam",
                exam_id=exam.id
            )
        )

    return render_template(
        "edit_exam.html",
        exam=exam
    )

@app.route(
    "/admin/question/edit/<int:question_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_question(question_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    question = Question.query.get_or_404(
        question_id
    )

    if request.method == "POST":

        question.question = request.form.get(
            "question"
        )

        question.option_a = request.form.get(
            "option_a"
        )

        question.option_b = request.form.get(
            "option_b"
        )

        question.option_c = request.form.get(
            "option_c"
        )

        question.option_d = request.form.get(
            "option_d"
        )

        question.answer = request.form.get(
            "answer"
        )

        db.session.commit()

        flash(
            "Question updated successfully",
            "success"
        )

        return redirect(
            url_for(
                "manage_exam",
                exam_id=question.exam_id
            )
        )

    return render_template(
        "edit_question.html",
        question=question
    )

@app.route(
    "/admin/question-bank/edit/<int:question_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_bank_question(question_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    question = QuestionBank.query.get_or_404(
        question_id
    )

    if request.method == "POST":

        question.subject = request.form.get(
            "subject"
        )

        question.topic = request.form.get(
            "topic"
        )

        question.question = request.form.get(
            "question"
        )

        question.option_a = request.form.get(
            "option_a"
        )

        question.option_b = request.form.get(
            "option_b"
        )

        question.option_c = request.form.get(
            "option_c"
        )

        question.option_d = request.form.get(
            "option_d"
        )

        question.answer = request.form.get(
            "answer"
        ).lower()

        db.session.commit()

        flash(
            "Question updated successfully",
            "success"
        )

        return redirect(
            url_for("question_bank")
        )

    return render_template(
        "edit_bank_question.html",
        question=question
    )

@app.route(
    "/admin/question-bank/delete/<int:question_id>"
)
@login_required
def delete_bank_question(question_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    question = QuestionBank.query.get_or_404(
        question_id
    )

    db.session.delete(question)

    db.session.commit()

    flash(
        "Question deleted successfully",
        "success"
    )

    return redirect(
        url_for("question_bank")
    )

@app.route(
    "/admin/question-bank/import/<int:question_id>"
)
@login_required
def select_exam_import(question_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    exams = Exam.query.order_by(
        Exam.title
    ).all()

    return render_template(
        "select_exam_import.html",
        exams=exams,
        question_id=question_id
    )
    
# Admin Question bank   
@app.route("/admin/question-bank")
@login_required
def question_bank():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "")
    subject = request.args.get("subject", "")

    query = QuestionBank.query

    if search:
        query = query.filter(
            QuestionBank.question.contains(search)
        )

    if subject:
        query = query.filter_by(subject=subject)

    pagination = query.order_by(
        QuestionBank.id.desc()
    ).paginate(
        page=page,
        per_page=10,
        error_out=False
    )

    subjects = db.session.query(
        QuestionBank.subject
    ).distinct().all()

    subjects = [s[0] for s in subjects]

    return render_template(
        "question_bank.html",
        pagination=pagination,
        subjects=subjects
    )
    
# ======================
# STUDENT EXAMS route
# ======================
@app.route("/exams")
@login_required
def exams():

    if current_user.role == "admin":
        return redirect(url_for("admin_dashboard"))

    exams = Exam.query.filter_by(
        is_published=True
    ).all()

    return render_template(
        "exams.html",
        exams=exams
    )

# ======================
# START EXAM route
# ======================
@app.route("/start-exam/<int:exam_id>")
@login_required
def start_exam(exam_id):

    exam = Exam.query.get_or_404(exam_id)

    if not exam.is_published:

        flash(
            "This exam is not available.",
            "warning"
        )

        return redirect(
            url_for("dashboard")
        )

    # Check exam date restrictions

    now = datetime.utcnow()

    if exam.start_date and now < exam.start_date:

        flash(
            "This exam has not started yet.",
            "warning"
        )

        return redirect(
            url_for("dashboard")
        )

    if exam.end_date and now > exam.end_date:

        flash(
            "This exam has ended.",
            "danger"
        )

        return redirect(
            url_for("dashboard")
        )

    # Multiple attempt control

    existing_session = ExamSession.query.filter_by(
        user_id=current_user.id,
        exam_id=exam.id,
        submitted=True
    ).first()

    if (
        existing_session
        and
        not exam.allow_multiple_attempts
    ):

        flash(
            "You have already taken this exam.",
            "warning"
        )

        return redirect(
            url_for("dashboard")
        )

    # Create new session

    session_token = str(uuid.uuid4())

    exam_session = ExamSession(
        user_id=current_user.id,
        exam_id=exam.id,
        session_token=session_token,
        start_time=datetime.utcnow(),
        last_activity=datetime.utcnow(),
        submitted=False
    )

    db.session.add(exam_session)
    db.session.commit()

    return redirect(
        url_for(
            "take_exam",
            exam_id=exam.id,
            session_token=session_token
        )
    )

#TAKE EXAM ROUTES
@app.route("/exam/<int:exam_id>")
@login_required
def take_exam(exam_id):

    exam = Exam.query.get_or_404(exam_id)

    questions = Question.query.filter_by(
        exam_id=exam.id
    ).all()

    if exam.shuffle_questions:
        random.shuffle(questions)

    session = ExamSession.query.filter_by(
        user_id=current_user.id,
        exam_id=exam.id,
        submitted=False
    ).order_by(
        ExamSession.id.desc()
    ).first()

    if not session:

        flash(
            "Exam session not found.",
            "danger"
        )

        return redirect(
            url_for("dashboard")
        )

    session.last_activity = datetime.utcnow()

    db.session.commit()

    return render_template(
        "take_exam.html",
        exam=exam,
        questions=questions,
        session=session
    )

#Submit exam route
@app.route("/submit-exam/<int:exam_id>", methods=["POST"])
@login_required
def submit_exam(exam_id):

    exam = Exam.query.get_or_404(exam_id)

    questions = Question.query.filter_by(
        exam_id=exam_id
    ).all()

    score = 0
    total = len(questions)

    for q in questions:

        selected_answer = request.form.get(
            str(q.id)
        )

        if selected_answer:

            selected_answer = (
                selected_answer
                .strip()
                .lower()
            )

            if (
                selected_answer
                ==
                q.answer.strip().lower()
            ):
                score += q.marks

    # ======================
    # SESSION HANDLING
    # ======================

    exam_session = ExamSession.query.filter_by(
        user_id=current_user.id,
        exam_id=exam_id
    ).first()

    if not exam_session:

        exam_session = ExamSession(
            user_id=current_user.id,
            exam_id=exam_id,
            start_time=datetime.utcnow(),
            session_token=str(uuid.uuid4())
        )

        db.session.add(exam_session)

    exam_session.score = score
    exam_session.submitted = True
    exam_session.last_activity = (
        datetime.utcnow()
    )

    db.session.commit()

    # ======================
    # RESULT CALCULATION
    # ======================

    total_marks = sum(
        q.marks for q in questions
    )

    percentage = (
        round(
            (score / total_marks) * 100,
            2
        )
        if total_marks > 0
        else 0
    )

    certificate_id = None

    # ======================
    # 🎓 CERTIFICATE LOGIC
    # ======================

    if percentage >= exam.pass_mark:

        file_path, cert_id, grade = (
            generate_certificate(
                current_user,
                exam,
                score,
                total
            )
        )

        cert = Certificate(
            user_id=current_user.id,
            exam_id=exam_id,
            score=score,
            total=total,
            grade=grade,
            certificate_id=cert_id
        )

        db.session.add(cert)
        db.session.commit()

        send_certificate_email(
            current_user,
            exam,
            file_path
        )

        certificate_id = cert_id

    return render_template(
        "results.html",
        session=exam_session,
        total=total,
        percentage=percentage,
        certificate=certificate_id
    )


@app.route("/result/<int:session_id>")
@login_required
def view_result(session_id):

    exam_session = ExamSession.query.get_or_404(session_id)

    total = exam_session.exam.total_marks or 100

    percentage = round(
        (exam_session.score / total) * 100,
        2
    )

    certificate = None

    if percentage >= exam_session.exam.pass_mark:
        certificate = f"CERT-{exam_session.id}"

    return render_template(
        "results.html",
        exam_session=exam_session,
        total=total,
        percentage=percentage,
        certificate=certificate
    )


@app.before_request
def anti_cheat_guard():

    if current_user.is_authenticated:

        # block exam tampering routes if needed
        blocked_routes = []

        if request.endpoint in blocked_routes:
            return redirect(url_for("dashboard"))


@app.route("/admin/results")
@login_required
def admin_results():

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    exams = (
        Exam.query
        .order_by(Exam.created_at.desc())
        .all()
    )

    return render_template(
        "admin_results.html",
        exams=exams
    )


@app.route("/admin/results/<int:exam_id>")
@login_required
def view_results(exam_id):

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    exam = Exam.query.get_or_404(exam_id)

    sessions = (
        ExamSession.query
        .filter_by(
            exam_id=exam_id,
            submitted=True
        )
        .order_by(ExamSession.score.desc())
        .all()
    )

    return render_template(
        "view_results.html",
        exam=exam,
        sessions=sessions
    )

@app.route("/results")
@login_required
def results():

    sessions = (
        ExamSession.query
        .filter_by(user_id=current_user.id)
        .order_by(ExamSession.start_time.desc())
        .all()
    )

    return render_template(
        "results.html",
        sessions=sessions
    )

@app.route("/admin/change-email", methods=["GET", "POST"])
@login_required
def change_email():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    if request.method == "POST":

        new_email = request.form.get("email")

        if not new_email:
            flash("Email cannot be empty", "danger")
            return redirect(url_for("change_email"))

        current_user.email = new_email
        db.session.commit()

        flash("Email updated successfully!", "success")
        return redirect(url_for("admin_dashboard"))

    return render_template("change_email.html")
  
  
import csv

from io import TextIOWrapper

from openpyxl import load_workbook

from docx import Document

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash
)

from flask_login import (
    login_required,
    current_user
)

from models import (
    db,
    Exam,
    Question
)


@app.route(
    "/admin/question-center/<int:exam_id>",
    methods=["GET", "POST"]
)
@login_required
def question_center(exam_id):

    if current_user.role != "admin":
        flash(
            "Access denied.",
            "danger"
        )
        return redirect(url_for("dashboard"))

    exam = Exam.query.get_or_404(exam_id)

    if request.method == "POST":

        file = request.files.get("file")

        if not file or file.filename == "":

            flash(
                "Please select a file.",
                "danger"
            )

            return redirect(request.url)

        filename = file.filename.lower()

        count = 0
        skipped = 0

        try:

            # ==================================
            # EXCEL FILE
            # ==================================
            if filename.endswith(".xlsx"):

                workbook = load_workbook(file)

                sheet = workbook.active

                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ):

                    try:

                        if not row:
                            skipped += 1
                            continue

                        if len(row) < 6:
                            skipped += 1
                            continue

                        if row[0] is None:
                            skipped += 1
                            continue

                        answer = str(
                            row[5]
                        ).strip().lower()

                        if answer not in [
                            "a",
                            "b",
                            "c",
                            "d"
                        ]:
                            skipped += 1
                            continue

                        question = Question(
                            exam_id=exam.id,
                            question=str(row[0]),
                            option_a=str(row[1]),
                            option_b=str(row[2]),
                            option_c=str(row[3]),
                            option_d=str(row[4]),
                            answer=answer,
                            marks=1
                        )

                        db.session.add(question)

                        count += 1

                    except Exception as e:

                        print(
                            f"Excel Row Error: {e}"
                        )

                        skipped += 1

            # ==================================
            # CSV FILE
            # ==================================
            elif filename.endswith(".csv"):

                csv_file = TextIOWrapper(
                    file.stream,
                    encoding="utf-8"
                )

                reader = csv.reader(csv_file)

                next(reader, None)

                for row in reader:

                    try:

                        if len(row) < 6:
                            skipped += 1
                            continue

                        answer = row[5].strip().lower()

                        if answer not in [
                            "a",
                            "b",
                            "c",
                            "d"
                        ]:
                            skipped += 1
                            continue

                        question = Question(
                            exam_id=exam.id,
                            question=row[0],
                            option_a=row[1],
                            option_b=row[2],
                            option_c=row[3],
                            option_d=row[4],
                            answer=answer,
                            marks=1
                        )

                        db.session.add(question)

                        count += 1

                    except Exception as e:

                        print(
                            f"CSV Row Error: {e}"
                        )

                        skipped += 1

            # ==================================
            # WORD FILE
            # ==================================
            elif filename.endswith(".docx"):

                doc = Document(file)

                flash(
                    "Word upload detected. DOCX parser not yet enabled.",
                    "info"
                )

                for para in doc.paragraphs:

                    text = para.text.strip()

                    if text:
                        print(text)

            # ==================================
            # INVALID FILE
            # ==================================
            else:

                flash(
                    "Only .xlsx, .csv and .docx files are supported.",
                    "danger"
                )

                return redirect(
                    request.url
                )

            db.session.commit()

            flash(
                f"{count} questions uploaded successfully. {skipped} skipped.",
                "success"
            )

            return redirect(
                url_for(
                    "manage_exam",
                    exam_id=exam.id
                )
            )

        except Exception as e:

            db.session.rollback()

            flash(
                f"Upload failed: {e}",
                "danger"
            )

            return redirect(
                request.url
            )

    return render_template(
        "question_center.html",
        exam=exam
    )


@app.route("/admin/save-draft/<int:exam_id>", methods=["POST"])
@login_required
def save_draft(exam_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    exam = Exam.query.get_or_404(exam_id)

    q = Question(
        exam_id=exam.id,
        question=request.form.get("question"),
        option_a=request.form.get("option_a"),
        option_b=request.form.get("option_b"),
        option_c=request.form.get("option_c"),
        option_d=request.form.get("option_d"),
        answer=request.form.get("answer"),
        status="draft"
    )

    db.session.add(q)
    db.session.commit()

    flash("Question saved as draft", "success")

    return redirect(url_for("manage_exam", exam_id=exam.id))
    

@app.route("/admin/publish-question/<int:question_id>")
@login_required
def publish_question(question_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    q = Question.query.get_or_404(question_id)

    q.status = "published"
    db.session.commit()

    flash("Question published", "success")

    return redirect(url_for("manage_exam", exam_id=q.exam_id))
  
  
@app.route("/admin/drafts/<int:exam_id>")
@login_required
def view_drafts(exam_id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    drafts = Question.query.filter_by(
        exam_id=exam_id,
        status="draft"
    ).all()

    return render_template("drafts.html", drafts=drafts)

@app.route("/result-pdf/<int:session_id>")
@login_required
def result_pdf(session_id):

    session = ExamSession.query.get_or_404(session_id)

    if session.user_id != current_user.id:
        abort(403)

    exam = Exam.query.get(session.exam_id)

    questions = Question.query.filter_by(
        exam_id=exam.id
    ).all()

    total_questions = len(questions)

    total_marks = sum(q.marks for q in questions)

    percentage = (
        round((session.score / total_marks) * 100, 2)
        if total_marks > 0 else 0
    )

    # =====================
    # GRADE
    # =====================

    if percentage >= 80:
        grade = "A"
        remark = "Excellent Performance"

    elif percentage >= 70:
        grade = "B"
        remark = "Very Good Performance"

    elif percentage >= 60:
        grade = "C"
        remark = "Good Performance"

    elif percentage >= 50:
        grade = "D"
        remark = "Fair Performance"

    else:
        grade = "F"
        remark = "Needs Improvement"

    status = (
        "PASS"
        if percentage >= exam.pass_mark
        else "FAIL"
    )

    # =====================
    # PDF
    # =====================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4
    )

    width, height = A4

    y = height - 50

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(
        width / 2,
        y,
        "SHALOM NEXUS CBT"
    )

    y -= 30

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawCentredString(
        width / 2,
        y,
        "OFFICIAL EXAMINATION RESULT"
    )

    y -= 50

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Candidate Information")

    y -= 25

    pdf.setFont("Helvetica", 11)

    pdf.drawString(
        60,
        y,
        f"Student Name: {current_user.full_name}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Exam Title: {exam.title}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Date Taken: {session.start_time.strftime('%d %B %Y')}"
    )

    y -= 40

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Performance Summary")

    y -= 25

    pdf.setFont("Helvetica", 11)

    pdf.drawString(
        60,
        y,
        f"Total Questions: {total_questions}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Score: {session.score}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Percentage: {percentage}%"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Grade: {grade}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Status: {status}"
    )

    y -= 20

    pdf.drawString(
        60,
        y,
        f"Remark: {remark}"
    )

    y -= 50

    result_id = (
        f"SNX-{session.id}-{session.user_id}"
    )

    pdf.setFont("Helvetica", 10)

    pdf.drawString(
        50,
        y,
        f"Result ID: {result_id}"
    )

    y -= 20

    pdf.drawString(
        50,
        y,
        "Generated by Shalom Nexus CBT System"
    )

    pdf.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Result_{exam.title}.pdf",
        mimetype="application/pdf"
    )

# ======================
# RUN
# ======================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)